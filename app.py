from flask import Flask, render_template, redirect, url_for, flash, request, session, current_app, jsonify, send_file, abort
import urllib
from werkzeug.security import generate_password_hash
from loguru import logger
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import joinedload
from functools import wraps
from openpyxl import load_workbook
from flask_login import LoginManager, UserMixin, login_required, current_user, login_user, logout_user
from extensions import db
from models import Employee
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect
from flask_migrate import Migrate


# 数据库错误处理装饰器
app = Flask(__name__)

# 初始化Flask-Migrate
migrate = Migrate(app, db)

def handle_db_error(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except SQLAlchemyError as e:
            db.session.rollback()
            logger.error(f'Database error: {str(e)}')
            flash(f'操作失败: 数据库错误', 'danger')
            with app.app_context():
                return redirect(url_for('admin_employee_list'))
        except Exception as e:
            db.session.rollback()
            logger.error(f'Unexpected error: {str(e)}')
            flash(f'操作失败: 系统错误', 'danger')
            with app.app_context():
                return redirect(url_for('admin_employee_list'))
    return wrapper
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login', next=request.url))
        if not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated_function


def readonly_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login', next=request.url))
        # 检查是否是管理员或者employee_id >= '2'
        if not (current_user.is_admin or current_user.employee_id >= '2'):
            abort(403)
        # 对于employee_id >= '2'的用户，检查是否是GET请求（只读操作）
        if current_user.employee_id >= '2' and request.method != 'GET':
            flash('您只能以只读模式访问管理后台', 'danger')
            return redirect(url_for('admin_dashboard'))
        return f(*args, **kwargs)
    return decorated_function
from models import EvaluationDimension as Dimension
from wtforms import StringField, SelectField, DecimalField, SubmitField, IntegerField, FloatField, FieldList, FormField
from wtforms.validators import DataRequired, Length, NumberRange, ValidationError
import pandas as pd
from datetime import datetime
import pytz
import os
import json
from io import BytesIO
from models import Employee, EvaluationDimension, EvaluationRecord, EvaluationScore, EvaluationTask
from forms import EvaluationForm, ScoreForm, LoginForm, ChangePasswordForm

# 配置应用
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'a1b2c3d4e5f6g7h8i9j0k1l2m3n4o5p6q7r8s9t0u1v2w3x4y5z6')
basedir = os.path.dirname(os.path.abspath(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'performance_evaluation.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)
import pytz
from datetime import datetime

@app.template_filter('shanghai_time')
def shanghai_time_filter(dt, format='%Y-%m-%d %H:%M:%S'):
    if dt is None:
        return ""
    # 将UTC时间转换为上海时区
    utc_dt = pytz.utc.localize(dt)
    shanghai_tz = pytz.timezone('Asia/Shanghai')
    return utc_dt.astimezone(shanghai_tz).strftime(format)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'a1b2c3d4e5f6g7h8i9j0k1l2m3n4o5p6q7r8s9t0u1v2w3x4y5z6')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'performance_evaluation.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False


# 初始化Flask-Login
login_manager = LoginManager(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return Employee.query.get(int(user_id))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = LoginForm()
    if form.validate_on_submit():
        employee = Employee.query.filter_by(employee_id=form.employee_id.data).first()
        if employee is None or not employee.check_password(form.password.data):
            flash('员工ID或密码不正确', 'danger')
            return redirect(url_for('login'))
        if employee.is_frozen:
            flash('该账户已被冻结，无法登录', 'danger')
            return redirect(url_for('login'))
        login_user(employee)
        next_page = request.args.get('next')
        return redirect(next_page or url_for('index'))
    return render_template('auth/login.html', form=form)

db.init_app(app)
csrf = CSRFProtect(app)

# 创建数据库表
from sqlalchemy import inspect
from sqlalchemy.orm import joinedload
with app.app_context():
    db.create_all()
    # 检查并添加缺失的表和列，使用try-except处理已存在的情况
    inspector = inspect(db.engine)
    
    # 处理evaluation_record表
    try:
        columns = inspector.get_columns('evaluation_record')
    except Exception as e:
        app.logger.error(f'处理evaluation_record表错误: {str(e)}')
        column_names = [col['name'] for col in columns]
        if 'task_id' not in column_names:
            try:
                db.engine.execute('ALTER TABLE evaluation_record ADD COLUMN task_id INTEGER')
                db.session.commit()
            except Exception as e:
                db.session.rollback()
      
        
        
    
    # 处理evaluation_task表
    try:
        task_columns = inspector.get_columns('evaluation_task')
        task_column_names = [col['name'] for col in task_columns]
        if 'status' not in task_column_names:
            db.session.execute(text('ALTER TABLE evaluation_task ADD COLUMN status VARCHAR(20) DEFAULT \'published\' NOT NULL'))
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.warning(f'处理evaluation_task表时出错: {str(e)}')
    

    
    # 处理evaluation_task表
    try:
        task_columns = inspector.get_columns('evaluation_task')
        task_column_names = [col['name'] for col in task_columns]
        if 'status' not in task_column_names:
                db.session.execute(text('ALTER TABLE evaluation_task ADD COLUMN status VARCHAR(20) DEFAULT \'published\' NOT NULL'))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.warning(f'处理evaluation_task表时出错: {str(e)}')
    
    # 处理evaluation_dimension表
    try:
        columns = inspector.get_columns('evaluation_dimension')
        column_names = [col['name'] for col in columns]
        if 'status' not in column_names:
            db.engine.execute('ALTER TABLE evaluation_dimension ADD COLUMN status VARCHAR(20) DEFAULT "published" NOT NULL')
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.warning(f'处理evaluation_dimension表时出错: {str(e)}')
    
    # 已移除处理user表的代码，系统实际使用employee表
    
    # 处理employee表
    try:
        columns = inspector.get_columns('employee')
        column_names = [col['name'] for col in columns]
        
        # 检查并添加employee_id字段
        if 'employee_id' not in column_names:
            db.engine.execute('ALTER TABLE employee ADD COLUMN employee_id VARCHAR(20)')
            db.engine.execute('UPDATE employee SET employee_id = CAST(id AS VARCHAR)')
            db.engine.execute('ALTER TABLE employee ALTER COLUMN employee_id SET NOT NULL')
            db.engine.execute('ALTER TABLE employee ADD CONSTRAINT employee_id_unique UNIQUE (employee_id)')
            db.session.commit()
        
        # 检查并添加password_hash字段
        if 'password_hash' not in column_names:
            db.engine.execute('ALTER TABLE employee ADD COLUMN password_hash VARCHAR(128)')
            db.session.commit()
        
        # 检查并添加is_admin字段
        if 'is_admin' not in column_names:
            db.engine.execute('ALTER TABLE employee ADD COLUMN is_admin BOOLEAN DEFAULT FALSE')
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.warning(f'处理employee表时出错: {str(e)}')
    
    # 设置默认密码'password'给所有没有密码的员工
    from models import Employee
    employees_without_password = Employee.query.filter(
        (Employee.password_hash.is_(None)) | 
        (Employee.password_hash == '') | 
        (Employee.password_hash == 'pbkdf2:sha256:150000$abc123$def456')
    ).all()
    for employee in employees_without_password:
        employee.set_password('password')
        db.session.add(employee)
    
    # 确保默认管理员账户存在
    default_admin = Employee.query.filter_by(employee_id='10000').first()
    if not default_admin:
        default_admin = Employee(
                employee_id='10000',  # 默认管理员ID（字符串类型）
                name='系统管理员',
                position='管理员',
                is_admin=True
            )
        db.session.add(default_admin)
    
    # 确保默认管理员密码和权限正确
    default_admin.set_password('password')
    default_admin.is_admin = True
    db.session.add(default_admin)
    
    db.session.commit()

# 评估者查询自己提交结果的路由
@app.route('/my_evaluations')
@login_required
def my_evaluations():
    # 获取当前用户关联的员工信息
    current_employee = current_user
    if not current_employee:
        flash('未找到关联的员工信息，请联系管理员', 'danger')
        return redirect(url_for('index'))
    
    # 获取任务筛选参数
    task_id = request.args.get('task_id')
    
    # 构建查询基础，关联任务信息
    from sqlalchemy.orm import joinedload
    # 查询已提交和已退回的评估记录
    query = EvaluationRecord.query.filter(
        EvaluationRecord.evaluator_id == current_employee.id,
        EvaluationRecord.status.in_(['submitted', 'returned', 'withdrawal_requested'])
    ).options(
        joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension),
        joinedload(EvaluationRecord.task)
    ).options(
        joinedload(EvaluationRecord.task),
        joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)
    )
    
    # 任务隔离：添加任务筛选条件
    if task_id:
        query = query.filter_by(task_id=task_id)
    
    # 执行查询
    evaluations = query.order_by(EvaluationRecord.submitted_at.desc()).all()

    # 按任务分组并检查是否有撤回请求
    task_withdrawal_status = {}
    for evaluation in evaluations:
        if evaluation.task_id not in task_withdrawal_status:
            task_withdrawal_status[evaluation.task_id] = {}
            task_withdrawal_status[evaluation.task_id]['has_withdrawal_requested'] = False
        if evaluation.status == 'withdrawal_requested':
            task_withdrawal_status[evaluation.task_id]['has_withdrawal_requested'] = True

    # 获取所有任务供筛选
    tasks = EvaluationTask.query.all()

    # 获取所有评估维度
    dimensions = EvaluationDimension.query.filter_by(status='published').order_by(EvaluationDimension.id).all()

    return render_template('evaluation/my_evaluations.html', evaluations=evaluations, tasks=tasks, selected_task_id=task_id, task_withdrawal_status=task_withdrawal_status, dimensions=dimensions)


@app.route('/request_withdrawal/<int:task_id>', methods=['POST'])
@login_required
def request_withdrawal(task_id):
    # 获取当前用户
    current_employee = current_user
    if not current_employee:
        flash('未找到关联的员工信息，请联系管理员', 'danger')
        return redirect(url_for('index'))
    
    # 查找该任务下当前用户的所有评估记录
    evaluations = EvaluationRecord.query.filter_by(
        task_id=task_id,
        evaluator_id=current_employee.id
    ).all()
    
    if not evaluations:
        flash('未找到该任务下的评估记录', 'danger')
        return redirect(url_for('my_evaluations'))
    
    # 检查所有评估记录状态是否都可以撤回
    for evaluation in evaluations:
        if evaluation.status != 'submitted':
            if evaluation.status == 'withdrawal_requested':
                flash('该任务已提交撤回申请，请勿重复提交', 'danger')
            else:
                flash('只有全部已提交的评估记录才能申请撤回', 'danger')
            return redirect(url_for('my_evaluations'))
    
    # 获取撤回原因
    reason = request.form.get('reason')
    if not reason or not reason.strip():
        flash('请填写撤回原因', 'danger')
        return redirect(url_for('my_evaluations'))
    
    # 将所有评估记录状态改为'withdrawal_requested'并记录撤回原因
    for evaluation in evaluations:
        evaluation.status = 'withdrawal_requested'
        evaluation.withdrawal_reason = reason
    db.session.commit()
    
    # 通知管理员（实际应用中可能需要发送消息或邮件）
    flash('撤回申请已提交，请等待管理员审核', 'success')
    
    return redirect(url_for('my_evaluations'))

# 主页路由
@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    form = ChangePasswordForm()
    if form.validate_on_submit():
        if not current_user.check_password(form.old_password.data):
            flash('旧密码不正确', 'danger')
            return redirect(url_for('change_password'))
        current_user.set_password(form.new_password.data)
        db.session.commit()
        flash('密码已成功更新', 'success')
        return redirect(url_for('index'))
    return render_template('auth/change_password.html', form=form)

@app.route('/')
@login_required
def index():
    return render_template('index.html')

# 管理页面路由
@app.route('/admin')
@login_required
@readonly_admin_required
def admin_dashboard():
    return render_template('admin/dashboard.html')

# 评估打分页面路由
@app.route('/evaluate', methods=['GET', 'POST'])
@login_required
def evaluate_page():
    # 管理员不能参与评估打分
    if current_user.employee_id == '10000':
        flash('管理员账号不能参与评估打分', 'danger')
        return redirect(url_for('index'))
    # 被冻结的员工不能参与评估打分
    if current_user.is_frozen:
        flash('该账户已被冻结，无法参与评估', 'danger')
        return redirect(url_for('index'))
    from models import Employee, EvaluationTask, EvaluationRecord
    from forms import EvaluationForm
    from sqlalchemy.orm import joinedload
    form = EvaluationForm()
    current_evaluator = current_user  # 当前登录用户作为评估者
    evaluatees = []
    evaluator_id = current_user.id  # 使用当前登录用户ID

    # 默认加载所有已发布的评估任务
    tasks = EvaluationTask.query.filter_by(status='published').all()
    selected_task = None
    if request.method == 'POST':
        if 'load_evaluatees' in request.form:
            # 获取除自己之外、管理员、被冻结的员工以及ID小于10001的员工
            evaluatees = Employee.query.filter(
    Employee.id != evaluator_id,
    Employee.employee_id != '10000',
    Employee.employee_id < '2',
    Employee.is_frozen == False
).all()
        # 处理任务选择
        task_id = request.form.get('task_id')
        if task_id:
            selected_task = EvaluationTask.query.get(task_id)

    # 查询当前用户的评估记录状态
    evaluations = EvaluationRecord.query.filter(
        EvaluationRecord.evaluator_id == evaluator_id,
        EvaluationRecord.status.in_(['submitted', 'returned', 'withdrawal_requested'])
    ).options(joinedload(EvaluationRecord.task)).all()

    # 按任务分组评估记录状态
    task_evaluation_status = {}
    for evaluation in evaluations:
        if evaluation.task_id not in task_evaluation_status:
            task_evaluation_status[evaluation.task_id] = {
                'has_submitted': False,
                'has_returned': False,
                'has_withdrawal_requested': False
            }
        if evaluation.status == 'submitted':
            task_evaluation_status[evaluation.task_id]['has_submitted'] = True
        elif evaluation.status == 'returned':
            task_evaluation_status[evaluation.task_id]['has_returned'] = True
        elif evaluation.status == 'withdrawal_requested':
            task_evaluation_status[evaluation.task_id]['has_withdrawal_requested'] = True

    return render_template('evaluation/evaluate.html', form=form, current_evaluator=current_evaluator, evaluatees=evaluatees, tasks=tasks, selected_task=selected_task, task_evaluation_status=task_evaluation_status)

# 获取评估表单
@app.route('/get_evaluation_form/<int:evaluatee_id>')
def get_evaluation_form(evaluatee_id):
    # 不能对管理员进行评估
    evaluatee = Employee.query.get_or_404(evaluatee_id)
    if evaluatee.employee_id == '10000':
        flash('不能对管理员账号进行评估', 'danger')
        return redirect(url_for('index'))
    # 获取被评估人信息
    evaluatee = Employee.query.get_or_404(evaluatee_id)
    # 查询所有评分维度
    dimensions = EvaluationDimension.query.all()
    # 如果没有维度，添加默认维度
    if not dimensions:
        default_dimensions = [
            {'name': '工作质量', 'description': '完成工作的准确性和质量', 'weight': 0.3},
            {'name': '工作效率', 'description': '完成工作的速度和效率', 'weight': 0.2},
            {'name': '团队合作', 'description': '与团队成员协作的能力', 'weight': 0.2},
            {'name': '创新性', 'description': '提出新想法和解决方案的能力', 'weight': 0.15},
            {'name': 'responsibility', 'description': '对工作的负责程度', 'weight': 0.15}
        ]
        for dim_data in default_dimensions:
            # 检查同名维度是否已存在
            existing_dim = EvaluationDimension.query.filter_by(name=dim_data['name']).first()
            if not existing_dim:
                dimension = EvaluationDimension(
                    name=dim_data['name'],
                    description=dim_data['description'],
                    weight=dim_data['weight']
                )
                db.session.add(dimension)
        db.session.commit()
        # 重新查询维度
        dimensions = EvaluationDimension.query.all()
    # 过滤掉ID为None的无效维度
    dimensions = [dim for dim in dimensions if dim.id is not None]
    # 创建评分表单
    score_form = EvaluationForm()
    # 设置最小条目数为0，防止自动添加空项
    score_form.scores.min_entries = 0
    for dimension in dimensions:
        if dimension.id is not None:
            score_form.scores.append_entry(data={'dimension_id': dimension.id})
        else:
              app.logger.warning(f'跳过ID为None的无效维度: {dimension.name}')
    return render_template('evaluation/evaluation_form.html', form=score_form, dimensions=dimensions, evaluatee=evaluatee)

# 提交评估
@app.route('/submit_evaluation', methods=['POST'])
@login_required
def submit_evaluation():
    # 管理员不能参与评估打分
    if current_user.employee_id == '10000':
        flash('管理员账号不能参与评估打分', 'danger')
        return redirect(url_for('index'))
    from models import db, Employee, EvaluationRecord, EvaluationScore, EvaluationDimension
    from forms import ScoreForm
    evaluator_id = current_user.id  # 使用当前登录用户ID作为评估者
    evaluatee_id = request.form.get('evaluatee_id')
    task_id = request.form.get('task_id')
    action = request.form.get('action', 'submit')

    if not evaluator_id or not evaluatee_id or not task_id:
        flash('参数错误', 'danger')
        return redirect(url_for('evaluate_page'))

    # 验证被评估者的employee_id < '2'
    evaluatee = Employee.query.get(evaluatee_id)
    if evaluatee and evaluatee.employee_id >= '2':
        flash('员工ID大于等于2的不作为被评估对象', 'danger')
        return redirect(url_for('evaluate_page'))

    # 获取所有维度
    dimensions = EvaluationDimension.query.all()
    dimension_ids = [d.id for d in dimensions]

    # 解析评分数据
    scores = {}
    valid = True
    for key, value in request.form.items():
        if key.startswith('scores-') and '-score' in key:
            try:
                index = int(key.split('-')[1])
                dimension_id = int(request.form.get(f'scores-{index}-dimension_id'))
                score = float(value)
                if dimension_id in dimension_ids and 1 <= score <= 5:
                    scores[dimension_id] = score
                else:
                    current_app.logger.error(f'无效的评分数据: dimension_id={dimension_id}, score={score}')
                    valid = False
            except Exception as e:
                current_app.logger.error(f'评分处理错误: {str(e)}')
                valid = False
                index = int(key.split('-')[1])
                dimension_id = int(request.form.get(f'scores-{index}-dimension_id'))
                score = float(value)
                if dimension_id in dimension_ids and 1 <= score <= 5:
                    scores[dimension_id] = score
                else:
                    current_app.logger.error(f'无效的评分数据: dimension_id={dimension_id}, score={score}')
                    valid = False
                    if not (dimension_id in dimension_ids and 1 <= score <= 5):
                            current_app.logger.error(f'无效的评分数据: dimension_id={dimension_id}, score={score}')
                            valid = False


    
    # 验证是否收集到所有维度的评分
    if len(scores) != len(dimensions):
        missing_dimensions = set(d.id for d in dimensions) - set(scores.keys())
        current_app.logger.error(f'缺少维度评分: {missing_dimensions}')
        valid = False

    # 正式提交时才验证所有维度都已打分，草稿保存不需要
    if action == 'submit' and (not valid or len(scores) != len(dimensions)):
        current_app.logger.error(f'Invalid submission: valid={valid}, scores_count={len(scores)}, dimensions_count={len(dimensions)}')
        flash('评分数据无效，请检查所有维度是否都已正确打分(1-5分)', 'danger')
        return redirect(url_for('evaluate_page', evaluator_id=evaluator_id))

    # 检查是否已提交该任务的评估
    existing_submission = EvaluationRecord.query.filter_by(
        evaluator_id=evaluator_id,
        evaluatee_id=evaluatee_id,
        task_id=task_id,
        status='submitted'
    ).options(joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)).first()
    
    if existing_submission:
        existing_submission.submitted_at = datetime.now(pytz.timezone('Asia/Shanghai'))
        existing_submission.status = 'submitted'  # 更新状态为已提交
        flash('你已提交该任务的评估，若需修改请通知管理员退回。', 'warning')
        return redirect(url_for('evaluate_page', evaluator_id=evaluator_id))
    
    # 查找或创建评估记录
    record = EvaluationRecord.query.filter_by(
        evaluator_id=evaluator_id,
        evaluatee_id=evaluatee_id,
        task_id=task_id
    ).options(joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)).first()
    
    # Find existing score for this record and dimension
    existing_score = EvaluationScore.query.filter_by(
    record_id=record.id if record else None,
    dimension_id=dimension_id
    ).first()

    if not record:
        record = EvaluationRecord(
                    evaluator_id=evaluator_id,
                    evaluatee_id=evaluatee_id,
                    task_id=task_id,
                    status='submitted',
                    submitted_at=datetime.now(pytz.timezone('Asia/Shanghai'))
                )
        db.session.add(record)

    # 更新分数
    for dimension_id, score in scores.items():
        score_record = EvaluationScore.query.filter_by(
                record_id=record.id,
                dimension_id=dimension_id
            ).first()

        if score_record:
            score_record.score = score
        else:
            score_record = EvaluationScore(
                record_id=record.id,
                dimension_id=dimension_id,
                score=score
            )
            db.session.add(score_record)

    # 更新状态
    if action == 'submit':
        record.status = 'submitted'
        record.submitted_at = datetime.utcnow()
        flash('评估已提交成功', 'success')
    else:
        record.status = 'submitted'
        record.task_id = task_id
        flash('评估已提交成功', 'success')

    db.session.commit()
    return redirect(url_for('evaluate_page', evaluator_id=evaluator_id))

@app.route('/submit_batch_evaluation', methods=['POST'])
def submit_batch_evaluation():
    # 初始化评分数据字典，确保在所有代码路径中都已定义
    scores_data = {}
    
    # 管理员不能参与评估打分
    if current_user.employee_id == '10000':
        flash('管理员账号不能参与评估打分', 'danger')
        return redirect(url_for('index'))
    # 检查当前用户是否被冻结
    if current_user.is_frozen:
        flash('该账户已被冻结，无法参与评估', 'danger')
        return redirect(url_for('index'))

    # 获取并验证评估者ID和任务ID
    try:
        evaluator_id = int(request.form.get('evaluator_id'))
        task_id = int(request.form.get('task_id'))
    except (ValueError, TypeError):
        flash('无效的评估者ID或任务ID', 'danger')
        return redirect(url_for('evaluate_page'))

    # 检查任务是否存在待审核的撤回申请
    has_pending_withdrawal = EvaluationRecord.query.filter_by(
        evaluator_id=evaluator_id,
        task_id=task_id,
        status='withdrawal_requested'
    ).first() is not None

    if has_pending_withdrawal:
        flash('该任务已提交撤回申请，待审核中，不能发起新的评估', 'danger')
        return redirect(url_for('batch_evaluate', evaluator_id=evaluator_id, task_id=task_id))

    # 已在前面验证过评估者ID和任务ID，此处无需重复验证

    action = request.form.get('action', 'submit')
    if not task_id:
        flash('任务ID不存在，无法提交评估', 'danger')
        return redirect(url_for('batch_evaluate', evaluator_id=evaluator_id))
    task_id = int(task_id)

    # 解析评分数据
    import re
    # 使用正则表达式匹配评分字段格式: scores[evaluatee_id][dimension_id]
    score_pattern = re.compile(r'scores\[(\d+)\]\[(\d+)\]')
    for key, value in request.form.items():
        match = score_pattern.search(key)
        if match:
            try:
                evaluatee_id = int(match.group(1))
                dimension_id = int(match.group(2))
                score_value = float(value)
                # 验证评分值在1-5范围内
                if 1 <= score_value <= 5:
                    # 将评分数据添加到scores_data
                    if evaluatee_id not in scores_data:
                        scores_data[evaluatee_id] = {}
                    scores_data[evaluatee_id][dimension_id] = score_value
                else:
                    current_app.logger.error(f'评分值 {score_value} 超出有效范围(1-5)，已跳过')
            except (ValueError, TypeError):
                current_app.logger.error(f'无效的评分数据: key={key}, value={value}')
                continue

    # 检查是否有评分数据
    if not scores_data:
        flash('没有要提交的评分数据', 'warning')
        return redirect(url_for('batch_evaluate', evaluator_id=evaluator_id, task_id=task_id))

    # 检查是否已提交该任务的评估
    existing_submissions = EvaluationRecord.query.filter_by(
        evaluator_id=evaluator_id,
        task_id=task_id,
        status='submitted'
    ).options(joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)).all()
    submitted_evaluatee_ids = [es.evaluatee_id for es in existing_submissions]

    # 检查是否有重复提交的被评估者
    duplicate_evaluatees = [eid for eid in scores_data.keys() if eid in submitted_evaluatee_ids]
    if duplicate_evaluatees:
        # 获取重复被评估者的姓名
        duplicate_employees = Employee.query.filter(Employee.id.in_(duplicate_evaluatees)).all()
        duplicate_names = [emp.name for emp in duplicate_employees]
        message = "以下被评估者的评估已提交：" + ', '.join(duplicate_names) + "，若需修改请通知管理员退回。"
        flash(message, 'warning')
        # 从评分数据中移除已提交的被评估者
        for eid in duplicate_evaluatees:
            if eid in scores_data:
                del scores_data[eid]
        # 如果没有剩余评分数据，重定向
        if not scores_data:
            return redirect(url_for('batch_evaluate', evaluator_id=evaluator_id, task_id=task_id))

    try:
        # 处理每个被评估者的评分
        if not scores_data:
            flash('没有要提交的评分数据', 'warning')
            return redirect(url_for('batch_evaluate', evaluator_id=evaluator_id, task_id=task_id))
        
        for evaluatee_id, dimensions in scores_data.items():
            # 获取或创建评估记录
            record = EvaluationRecord.query.filter_by(
                evaluator_id=evaluator_id,
                evaluatee_id=evaluatee_id,
                task_id=task_id
            ).first()
            if not record:
                record = EvaluationRecord(
                    evaluator_id=evaluator_id,
                    evaluatee_id=evaluatee_id,
                    task_id=task_id,
                    status='draft'
                )
                db.session.add(record)
            
            # 更新评分
            for dimension_id, score_value in dimensions.items():
                dimension = EvaluationDimension.query.get(dimension_id)
                if dimension:
                    score_record = EvaluationScore.query.filter_by(
                        record_id=record.id,
                        dimension_id=dimension_id
                    ).first()
                    if not score_record:
                        score_record = EvaluationScore(
                            record_id=record.id,
                            dimension_id=dimension_id,
                            score=score_value
                        )
                        db.session.add(score_record)
                    else:
                        score_record.score = score_value

            # 更新当前评估记录状态
            record.status = 'submitted'
            if action == 'submit':
                record.submitted_at = datetime.utcnow()
            record.updated_at = datetime.utcnow()

        db.session.commit()
        flash('批量评估提交成功！', 'success')
        return redirect(url_for('batch_evaluate', evaluator_id=evaluator_id))
    except Exception as e:
        db.session.rollback()
        flash(f'提交失败: {str(e)}', 'danger')
    finally:
        db.session.close()

    return redirect(url_for('batch_evaluate', evaluator_id=evaluator_id, task_id=task_id))

# 评估任务管理路由
@app.route('/admin/tasks')
@login_required
@readonly_admin_required
def admin_task_list():
    from models import EvaluationTask
    tasks = EvaluationTask.query.all()
    return render_template('admin/task_list.html', tasks=tasks)

@app.route('/admin/tasks/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_task():
    # 排除管理员账号(10000)被添加到评估任务
    employees = Employee.query.filter(Employee.employee_id != '10000').all()
    from forms import EvaluationTaskForm
    from models import EvaluationTask
    form = EvaluationTaskForm()
    if form.validate_on_submit():
        task_name = f"信息技术部20{form.year.data}年{form.quarter.data}季度绩效互评"
        # 检查任务是否已存在
        existing_task = EvaluationTask.query.filter_by(
            year=2000 + int(form.year.data),
            quarter=form.quarter.data
        ).first()
        if existing_task:
            flash('该季度评估任务已存在', 'danger')
            return redirect(url_for('admin_add_task'))
        
        task = EvaluationTask(
            year=form.year.data,
            quarter=form.quarter.data,
            name=task_name,
            status='published'
        )
        db.session.add(task)
        db.session.commit()
        flash('任务发布成功', 'success')
        return redirect(url_for('admin_task_list'))
    return render_template('admin/task_form.html', form=form, employees=employees)

@app.route('/admin/tasks/clear/<int:task_id>')
@admin_required
def admin_clear_task_data(task_id):
    from models import db, EvaluationRecord
    from sqlalchemy import func, EvaluationScore
    try:
        # 先删除关联的评分记录
        EvaluationScore.query.filter(EvaluationScore.record_id.in_(
            db.session.query(EvaluationRecord.id).filter_by(task_id=task_id)
        )).delete(synchronize_session=False)
        # 再删除评估记录
        EvaluationRecord.query.filter_by(task_id=task_id).delete(synchronize_session=False)
        db.session.commit()
        flash('任务评估数据已成功清除', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'清除失败: {str(e)}', 'danger')
    return redirect(url_for('admin_task_list'))

@app.route('/admin/tasks/delete/<int:id>')
@login_required
@admin_required
def admin_delete_task(id):
    from models import db, EvaluationTask
    task = EvaluationTask.query.get_or_404(id)
    try:
        db.session.delete(task)
        db.session.commit()
        flash('评估任务及其所有相关记录已成功删除', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败: {str(e)}', 'danger')
    return redirect(url_for('admin_task_list'))

# 员工管理路由
@app.route('/admin/employees/reset-password/<string:employee_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_employee_reset_password(employee_id):
    from models import db, Employee
    from forms import EmployeeForm
    import logging
    from flask import request
    logging.basicConfig(level=logging.DEBUG)
    logger = logging.getLogger(__name__)
    
    employee = Employee.query.filter_by(employee_id=employee_id).first_or_404()
    logger.debug(f'找到员工: {employee.name} (ID: {employee.employee_id})')
    
    # 创建空表单，手动设置字段值
    form = EmployeeForm(edit_id=employee.id, reset_password=True)
    form.employee_id.data = employee.employee_id
    form.employee_id.render_kw = {'readonly': True}
    form.name.data = employee.name
    form.name.render_kw = {'readonly': True}
    form.position.data = employee.position
    form.position.render_kw = {'readonly': True}
    
    if request.method == 'POST':
        logger.debug(f'表单提交数据: {request.form}')
        if form.validate():
            logger.debug('表单验证通过')
            try:
                if form.password.data:
                    logger.debug(f'密码长度: {len(form.password.data)}')
                    # 直接设置密码
                    employee.password_hash = generate_password_hash(form.password.data)
                    db.session.commit()
                    logger.debug('密码重置成功，已提交数据库')
                    flash('员工密码重置成功', 'success')
                    return redirect(url_for('admin_employee_list'))
                else:
                    logger.debug('密码为空')
                    flash('密码不能为空', 'danger')
            except Exception as e:
                db.session.rollback()
                logger.error(f'重置失败: {str(e)}')
                flash(f'重置失败: {str(e)}', 'danger')
        else:
            logger.debug(f'表单验证失败: {form.errors}')
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f'{getattr(form, field).label.text}：{error}', 'danger')
    return render_template('admin/employee_form.html', form=form, employee=employee, reset_password=True)

@app.route('/admin/employees')
@login_required
@readonly_admin_required
def admin_employee_list():
    from models import Employee
    from forms import EmployeeForm
    # 排除管理员账号(10000)参与评估
    employees = Employee.query.filter(Employee.employee_id != '10000').all()
    from forms import EmployeeForm
    return render_template('admin/employee_list.html', employees=employees, form=EmployeeForm())

@app.route('/admin/employees/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_employee_add():
    from models import db, Employee
    from forms import EmployeeForm
    form = EmployeeForm()
    if form.validate_on_submit():
        employee = Employee(
            employee_id=form.employee_id.data,
            name=form.name.data,
            position=form.position.data
        )
        employee.set_password(form.password.data)
        db.session.add(employee)
        try:
            db.session.commit()
            flash('员工信息添加成功', 'success')
            return redirect(url_for('admin_employee_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败: {str(e)}', 'danger')
    return render_template('admin/employee_form.html', form=form)

@app.route('/admin/employees/edit/<string:employee_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_employee_edit(employee_id):
    from models import Employee
    from forms import EmployeeForm
    employee = Employee.query.filter_by(employee_id=employee_id).first_or_404()
    form = EmployeeForm(obj=employee, edit_id=employee.id)
    if form.validate_on_submit():
            try:
                # 只更新职位字段
                employee.position = form.position.data
                # 如果密码不为空，则更新密码
                if form.password.data:
                    employee.set_password(form.password.data)
                db.session.commit()
                flash('员工信息更新成功', 'success')
                # 始终返回员工列表页面
                return redirect(url_for('admin_employee_list'))
            except Exception as e:
                db.session.rollback()
                flash(f'更新失败: {str(e)}', 'danger')
    return render_template('admin/employee_form.html', form=form, employee=employee)

@app.route('/admin/employees/toggle-frozen/<string:employee_id>')
@login_required
@admin_required
def admin_employee_toggle_frozen(employee_id):
    from models import db, Employee
    employee = Employee.query.filter_by(employee_id=employee_id).first_or_404()
    try:
        employee.is_frozen = not employee.is_frozen
        db.session.commit()
        if employee.is_frozen:
            flash(f'员工 {employee.name} 已成功冻结', 'success')
        else:
            flash(f'员工 {employee.name} 已成功解冻', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'操作失败: {str(e)}', 'danger')
    return redirect(url_for('admin_employee_list'))

@app.route('/admin/employees/delete/<int:id>')
@login_required
@admin_required
def admin_employee_delete(id):
    from models import db, Employee, EvaluationRecord
    employee = Employee.query.get_or_404(id)
    try:
        # 删除相关的评估记录
        EvaluationRecord.query.filter(
            (EvaluationRecord.evaluator_id == id) | (EvaluationRecord.evaluatee_id == id)
        ).delete()
        db.session.delete(employee)
        db.session.commit()
        flash('员工信息已删除', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败: {str(e)}', 'danger')
    return redirect(url_for('admin_employee_list'))

# 评估维度管理路由
@app.route('/admin/dimensions')
@login_required
@readonly_admin_required
def admin_dimension_list():
    from models import EvaluationDimension
    from forms import DimensionForm
    dimensions = EvaluationDimension.query.all()
    total_weight = sum(d.weight for d in dimensions)
    return render_template('admin/dimension_list.html', dimensions=dimensions, total_weight=total_weight)

@app.route('/admin/dimensions/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_dimension_add():
    from models import db, EvaluationDimension
    from forms import DimensionForm
    form = DimensionForm()
    current_dimensions = EvaluationDimension.query.all()
    total_weight = sum(d.weight for d in current_dimensions)

    if form.validate_on_submit():
        # 检查总权重
        new_total = total_weight + float(form.weight.data)
        if new_total > 1.0001:
            flash('添加后总权重超过100%，请调整权重值', 'danger')
            return render_template('admin/dimension_form.html', form=form, total_weight=total_weight)

        dimension = EvaluationDimension(
            name=form.name.data,
            weight=form.weight.data,
            status='published'
        )
        db.session.add(dimension)
        db.session.commit()
        flash('评估维度添加成功', 'success')
        return redirect(url_for('admin_dimension_list'))

    return render_template('admin/dimension_form.html', form=form, total_weight=total_weight)

@app.route('/admin/dimensions/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_dimension_edit(id):
    from models import EvaluationDimension
    from forms import DimensionForm
    dimension = EvaluationDimension.query.get_or_404(id)
    form = DimensionForm(obj=dimension)
    current_dimensions = EvaluationDimension.query.filter(EvaluationDimension.id != id).all()
    total_weight = sum(d.weight for d in current_dimensions)

    if form.validate_on_submit():
        # 检查总权重
        new_total = total_weight + float(form.weight.data)
        if new_total > 1.0001:
            flash('修改后总权重超过100%，请调整权重值', 'danger')
            return render_template('admin/dimension_form.html', form=form, dimension=dimension, total_weight=total_weight)

        dimension.name = form.name.data
        dimension.weight = form.weight.data
        db.session.commit()
        flash('评估维度更新成功', 'success')
        return redirect(url_for('admin_dimension_list'))

    return render_template('admin/dimension_form.html', form=form, dimension=dimension, total_weight=total_weight)

@app.route('/admin/dimensions/delete/<int:id>')
@login_required
@admin_required
def admin_dimension_delete(id):
    from models import db, EvaluationDimension, EvaluationScore
    dimension = EvaluationDimension.query.get_or_404(id)
    try:
        # 删除相关的评分记录
        EvaluationScore.query.filter_by(dimension_id=id).delete()
        db.session.delete(dimension)
        db.session.commit()
        flash('评估维度已删除', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败: {str(e)}', 'danger')
    return redirect(url_for('admin_dimension_list'))

def generate_evaluation_summary(evaluations, evaluators, evaluatees):
    """生成评估汇总统计数据的公共函数
    Args:
        evaluations: 评估记录列表
        evaluators: 评估者列表
        evaluatees: 被评估者列表
    Returns:
        summary_data: 评估矩阵数据
        averages: 被评估者平均分
    """
    summary_data = {}
    averages = {}
    evaluatee_scores = {}

    # 初始化数据结构
    for evaluatee in evaluatees:
        summary_data[evaluatee.id] = {}
        averages[evaluatee.id] = 0
        for evaluator in evaluators:
            summary_data[evaluatee.id][evaluator.id] = '-'  # 默认值

    # 填充数据
    for eval_rec in evaluations:
        # 检查被评估者ID是否有效
        if eval_rec.evaluatee_id not in summary_data:
            current_app.logger.warning(f"无效的被评估者ID: {eval_rec.evaluatee_id}，评估记录ID: {eval_rec.id}")
            continue
        
        if eval_rec.evaluatee_id not in evaluatee_scores:
            evaluatee_scores[eval_rec.evaluatee_id] = []
        evaluatee_scores[eval_rec.evaluatee_id].append(eval_rec.total_score)
        
        # 检查评估者ID是否有效
        if eval_rec.evaluator_id in summary_data[eval_rec.evaluatee_id]:
            summary_data[eval_rec.evaluatee_id][eval_rec.evaluator_id] = eval_rec.total_score
        else:
            current_app.logger.warning(f"无效的评估者ID: {eval_rec.evaluator_id}，评估记录ID: {eval_rec.id}")

    # 如果没有有效的评分数据，返回空字典
    if not evaluatee_scores:
        return {}, {}

    # 计算平均分
    for evaluatee_id, scores in evaluatee_scores.items():
        if scores:
            averages[evaluatee_id] = sum(scores) / len(scores)

    return summary_data, averages

# 评估查询路由
@app.route('/admin/evaluations')
@login_required
@readonly_admin_required
def admin_evaluation_query():
    from models import Employee, EvaluationRecord, EvaluationTask
    from openpyxl.styles import Font
    from openpyxl.styles import Font
    from forms import EvaluationSearchForm
    evaluatee_id = request.args.get('evaluatee_id')
    task_id = request.args.get('task_id')
    evaluator_id = request.args.get('evaluator_id')

    # 获取所有员工和任务用于筛选
    employees = Employee.query.filter(Employee.employee_id != '10000').all()
    tasks = EvaluationTask.query.filter_by(status='published').all()

    # 构建查询，预加载关联数据
    # 查询所有已提交的评估记录
    from sqlalchemy.orm import joinedload
    from sqlalchemy import func, and_
    # 排除管理员相关的评估记录
    admin_employee = Employee.query.filter_by(employee_id='10000').first()
    query = EvaluationRecord.query.filter_by(status='submitted').options(joinedload(EvaluationRecord.task), joinedload(EvaluationRecord.evaluator))
    if admin_employee:
        query = query.filter(
            ~EvaluationRecord.evaluator_id.in_([admin_employee.id]) &
            ~EvaluationRecord.evaluatee_id.in_([admin_employee.id])
        )
    # 包含所有评估对象（排除管理员）
    # 排除employee_id >= '2'的评估者
    non_valid_evaluators = Employee.query.filter(Employee.employee_id >= '2').all()
    non_valid_evaluator_ids = [emp.id for emp in non_valid_evaluators]
    if non_valid_evaluator_ids:
        query = query.filter(~EvaluationRecord.evaluator_id.in_(non_valid_evaluator_ids))
    # 排除employee_id >= '2'的被评估者
    non_valid_evaluatees = Employee.query.filter(Employee.employee_id >= '2').all()
    non_valid_evaluatee_ids = [emp.id for emp in non_valid_evaluatees]
    if non_valid_evaluatee_ids:
        query = query.filter(~EvaluationRecord.evaluatee_id.in_(non_valid_evaluatee_ids))
        query = query.filter(~EvaluationRecord.evaluatee_id.in_(non_valid_evaluatee_ids))
    if task_id:
        query = query.filter_by(task_id=task_id)
    
    # 添加去重逻辑：每个评估者-被评估者-任务组合只保留最新记录
    subquery = query.with_entities(
        EvaluationRecord.evaluator_id,
        EvaluationRecord.evaluatee_id,
        EvaluationRecord.task_id,
        func.max(EvaluationRecord.id).label('max_id')
    ).group_by(
        EvaluationRecord.evaluator_id,
        EvaluationRecord.evaluatee_id,
        EvaluationRecord.task_id
    ).subquery()
    
    query = query.join(
        subquery,
        and_(
            EvaluationRecord.evaluator_id == subquery.c.evaluator_id,
            EvaluationRecord.evaluatee_id == subquery.c.evaluatee_id,
            EvaluationRecord.task_id == subquery.c.task_id,
            EvaluationRecord.id == subquery.c.max_id
        )
    )
    
    query = query.options(
        joinedload(EvaluationRecord.task),
        joinedload(EvaluationRecord.evaluator),
        joinedload(EvaluationRecord.evaluatee),
        joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)
    )
    if evaluatee_id:
        query = query.filter_by(evaluatee_id=evaluatee_id)
    if evaluator_id:
        query = query.filter_by(evaluator_id=evaluator_id)
    # 强制任务隔离：必须提供task_id才能查询
    if task_id:
        query = query.filter_by(task_id=task_id)
    evaluations = query.order_by(EvaluationRecord.submitted_at.desc()).all()
    # 按task_id排序确保groupby能正确分组
    evaluations.sort(key=lambda x: x.task_id)
    flash(f'找到 {len(evaluations)} 条已提交的评估记录', 'info')

    # 获取评估者和被评估者列表（排除管理员和employee_id >= '2'的人员）
    evaluators = Employee.query.filter(Employee.employee_id != '10000', Employee.employee_id < '2').all()
    evaluatees = Employee.query.filter(Employee.employee_id != '10000', Employee.employee_id < '2').all()
    
    # 生成汇总统计数据
    from itertools import groupby
    task_summaries = {}
    # 按任务分组评估记录
    for task_id, task_evaluations in groupby(evaluations, key=lambda x: x.task_id):
        task_evaluations = list(task_evaluations)
        # 初始化该任务的汇总数据
        summary_data = {e.id: {} for e in evaluatees}
        averages = {e.id: 0 for e in evaluatees}
        
        # 填充评分数据
        for evaluation in task_evaluations:
            # 排除管理员和employee_id >= '2'的人员已在查询中处理
            if evaluation.evaluator_id in [e.id for e in evaluators] and evaluation.evaluatee_id in [e.id for e in evaluatees]:

                summary_data[evaluation.evaluatee_id][evaluation.evaluator_id] = round(evaluation.total_score, 2)
        
        # 计算平均分
        for evaluatee_id in summary_data:
            scores = list(summary_data[evaluatee_id].values())
            if scores:
                averages[evaluatee_id] = round(sum(scores) / len(scores), 2)
            else:
                averages[evaluatee_id] = 0
        
        # 保存该任务的汇总数据
        task_summaries[task_id] = {
            'summary_data': summary_data,
            'averages': averages
        }

    return render_template(
        'admin/evaluation_query.html', 
        evaluations=evaluations, 
        employees=employees, 
        evaluators=evaluators, 
        evaluatees=evaluatees, 
        tasks=tasks,
        task_summaries=task_summaries
    )



# 领导评估统计功能
@app.route('/admin/leader_evaluation_stats')
@login_required
@admin_required
def admin_leader_evaluation_stats():
    from models import Employee, EvaluationRecord, EvaluationTask
    from openpyxl.styles import Font
    from forms import EvaluationSearchForm
    evaluatee_id = request.args.get('evaluatee_id')
    task_id = request.args.get('task_id')
    evaluator_id = request.args.get('evaluator_id')

    # 获取所有员工和任务用于筛选
    employees = Employee.query.filter(Employee.employee_id != '10000').all()
    tasks = EvaluationTask.query.filter_by(status='published').all()

    # 构建查询，预加载关联数据
    # 查询所有已提交的评估记录
    from sqlalchemy.orm import joinedload
    from sqlalchemy import func, and_
    # 排除管理员相关的评估记录
    admin_employee = Employee.query.filter_by(employee_id='10000').first()
    query = EvaluationRecord.query.filter_by(status='submitted').options(joinedload(EvaluationRecord.task), joinedload(EvaluationRecord.evaluator))
    if admin_employee:
        query = query.filter(
            ~EvaluationRecord.evaluator_id.in_([admin_employee.id]) &
            ~EvaluationRecord.evaluatee_id.in_([admin_employee.id])
        )
    # 只包含employee_id >= '2'的评估者提交的记录
    valid_evaluators = Employee.query.filter(Employee.employee_id >= '2').all()
    valid_evaluator_ids = [emp.id for emp in valid_evaluators]
    if valid_evaluator_ids:
        query = query.filter(EvaluationRecord.evaluator_id.in_(valid_evaluator_ids))
    if task_id:
        query = query.filter_by(task_id=task_id)
    
    # 添加去重逻辑：每个评估者-被评估者-任务组合只保留最新记录
    subquery = query.with_entities(
        EvaluationRecord.evaluator_id,
        EvaluationRecord.evaluatee_id,
        EvaluationRecord.task_id,
        func.max(EvaluationRecord.id).label('max_id')
    ).group_by(
        EvaluationRecord.evaluator_id,
        EvaluationRecord.evaluatee_id,
        EvaluationRecord.task_id
    ).subquery()
    
    query = query.join(
        subquery,
        and_(
            EvaluationRecord.evaluator_id == subquery.c.evaluator_id,
            EvaluationRecord.evaluatee_id == subquery.c.evaluatee_id,
            EvaluationRecord.task_id == subquery.c.task_id,
            EvaluationRecord.id == subquery.c.max_id
        )
    )
    
    query = query.options(
        joinedload(EvaluationRecord.task),
        joinedload(EvaluationRecord.evaluator),
        joinedload(EvaluationRecord.evaluatee),
        joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)
    )
    if evaluatee_id:
        query = query.filter_by(evaluatee_id=evaluatee_id)
    if evaluator_id:
        query = query.filter_by(evaluator_id=evaluator_id)
    # 强制任务隔离：必须提供task_id才能查询
    if task_id:
        query = query.filter_by(task_id=task_id)
    evaluations = query.order_by(EvaluationRecord.submitted_at.desc()).all()
    # 按task_id排序确保groupby能正确分组
    evaluations.sort(key=lambda x: x.task_id)
    flash(f'找到 {len(evaluations)} 条已提交的领导评估记录', 'info')

    # 获取评估者和被评估者列表
    # 只包含employee_id >= '2'的评估者
    evaluators = Employee.query.filter(Employee.employee_id != '10000', Employee.employee_id >= '2').all()
    evaluatees = Employee.query.filter(Employee.employee_id != '10000', Employee.employee_id < '2').all()

    # 生成任务汇总数据
    from collections import defaultdict
    from itertools import groupby
    # 假设generate_evaluation_summary函数在当前作用域可用
    task_summaries = defaultdict(dict)
    # 按任务分组评估记录
    for task_id, task_evaluations in groupby(evaluations, key=lambda x: x.task_id):
        task_evaluations_list = list(task_evaluations)
        # 调用公共汇总函数生成统计数据
        summary_data, averages = generate_evaluation_summary(task_evaluations_list, evaluators, evaluatees)
        # 保存汇总数据
        task_summaries[task_id] = {
            'summary_data': summary_data,
            'averages': averages
        }

    # 评估者和被评估者列表已在上方定义
    
    return render_template(
        'admin/leader_evaluation_stats.html', 
        evaluations=evaluations, 
        employees=employees, 
        tasks=tasks, 
        evaluators=evaluators, 
        evaluatees=evaluatees, 
        task_summaries=task_summaries
    )


# 导出领导评估汇总表
@app.route('/admin/leader_evaluations/export')
@login_required
@admin_required
def export_leader_evaluation_summary():
    from flask import send_file
    from models import Employee, EvaluationRecord, EvaluationTask
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    # 获取所有已提交的评估
    task_id = request.args.get('task_id', type=int)
    query = EvaluationRecord.query.filter_by(status='submitted')
    if task_id:
        query = query.filter_by(task_id=task_id)
    # 排除管理员相关的评估记录
    admin_employee = Employee.query.filter_by(employee_id='10000').first()
    if admin_employee:
        query = query.filter(
            ~EvaluationRecord.evaluator_id.in_([admin_employee.id]) &
            ~EvaluationRecord.evaluatee_id.in_([admin_employee.id])
        )
    # 只包含employee_id >= '2'的评估者
    valid_evaluators = Employee.query.filter(Employee.employee_id >= '2').all()
    valid_evaluator_ids = [emp.id for emp in valid_evaluators]
    if valid_evaluator_ids:
        query = query.filter(EvaluationRecord.evaluator_id.in_(valid_evaluator_ids))
    all_evaluations = query.options(joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)).all()
    
    # 获取任务名称用于文件名
    task = None
    task_name = "all_tasks"
    if task_id:
        task = EvaluationTask.query.get(task_id)
        if task:
            task_name = task.name.replace(" ", "_")
    evaluators = Employee.query.filter(Employee.employee_id != '10000', Employee.employee_id >= '2').all()
    evaluatees = Employee.query.filter(Employee.employee_id != '10000', Employee.employee_id < '2').all()

    # 调用公共汇总函数生成统计数据
    summary_data, averages = generate_evaluation_summary(all_evaluations, evaluators, evaluatees)

    # 计算平均分（保留两位小数用于导出）
    averages = {eid: round(score, 2) for eid, score in averages.items()}

    # 创建Excel文件
    df_data = []
    for evaluatee in evaluatees:
        row = {'评估对象': evaluatee.name}
        for evaluator in evaluators:
            row[evaluator.name] = summary_data[evaluatee.id][evaluator.id]
        row['平均分'] = averages.get(evaluatee.id, '-')
        df_data.append(row)

    df = pd.DataFrame(df_data)

    # 保存到内存
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 添加表头
        df.to_excel(writer, index=False, sheet_name='领导评估汇总', startrow=1)  # 预留第一行给表头
        worksheet = writer.sheets['领导评估汇总']
        
        # 设置表头文本
        if task:
            title = f"信息技术部{task.year}年{task.quarter}季度领导绩效评估汇总"
        else:
            title = "信息技术部领导绩效评估汇总"
        
        # 创建表头单元格并设置样式
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = worksheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        
        # 设置列宽和单元格样式
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # 输出Excel文件
    output.seek(0)
    filename = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)


# 导出评估汇总表
@app.route('/admin/evaluations/export')
@login_required
@admin_required
def export_evaluation_summary():
    from models import Employee, EvaluationRecord, EvaluationTask
    from openpyxl.styles import Font
    # 获取所有已提交的评估
    task_id = request.args.get('task_id', type=int)
    query = EvaluationRecord.query.filter_by(status='submitted')
    if task_id:
        query = query.filter_by(task_id=task_id)
    # 排除管理员相关的评估记录
    admin_employee = Employee.query.filter_by(employee_id='10000').first()
    if admin_employee:
        query = query.filter(
            ~EvaluationRecord.evaluator_id.in_([admin_employee.id]) &
            ~EvaluationRecord.evaluatee_id.in_([admin_employee.id])
        )
    # 排除employee_id >= '2'的评估者
    non_valid_evaluators = Employee.query.filter(Employee.employee_id >= '2').all()
    non_valid_evaluator_ids = [emp.id for emp in non_valid_evaluators]
    if non_valid_evaluator_ids:
        query = query.filter(~EvaluationRecord.evaluator_id.in_(non_valid_evaluator_ids))
    all_evaluations = query.options(joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)).all()
    
    # 获取任务名称用于文件名
    task = None
    task_name = "all_tasks"
    if task_id:
        task = EvaluationTask.query.get(task_id)
        if task:
            task_name = task.name.replace(" ", "_")
    evaluators = Employee.query.filter(Employee.employee_id != '10000', Employee.employee_id < '2').all()
    evaluatees = Employee.query.filter(Employee.employee_id != '10000', Employee.employee_id < '2').all()

    # 调用公共汇总函数生成统计数据
    summary_data, averages = generate_evaluation_summary(all_evaluations, evaluators, evaluatees)

    # 计算平均分（保留两位小数用于导出）
    averages = {eid: round(score, 2) for eid, score in averages.items()}

    # 创建Excel文件
    df_data = []
    for evaluatee in evaluatees:
        row = {'评估对象': evaluatee.name}
        for evaluator in evaluators:
            row[evaluator.name] = summary_data[evaluatee.id][evaluator.id]
        row['平均分'] = averages.get(evaluatee.id, '-')
        df_data.append(row)

    df = pd.DataFrame(df_data)

    # 保存到内存
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 添加表头
        df.to_excel(writer, index=False, sheet_name='评估汇总', startrow=1)  # 预留第一行给表头
        worksheet = writer.sheets['评估汇总']
        
        # 设置表头文本
        if task:
            title = f"信息技术部{task.year}年{task.quarter}季度绩效互评汇总"
        else:
            title = "信息技术部绩效互评汇总"
        
        # 合并单元格并设置表头
        worksheet.merge_cells('A1:Z1')  # 合并A1到Z1的单元格
        worksheet['A1'] = title
        
        # 设置表头样式
        header_font = Font(bold=True, size=14)
        worksheet['A1'].font = header_font
    output.seek(0)

    # 发送文件
    return send_file(output, as_attachment=True, download_name=f'绩效评估汇总表_{task_name}_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# 查看评估详情
@app.route('/admin/evaluations/view/<int:id>')
@login_required
@admin_required
def view_evaluation(id):
    from models import EvaluationRecord
    evaluation = EvaluationRecord.query.options(joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)).get_or_404(id)
    return render_template('admin/evaluation_view.html', evaluation=evaluation)

# 退回单个评估
@app.route('/admin/evaluations/withdrawal_requests')
@login_required
@admin_required
def withdrawal_requests():
    from models import EvaluationRecord
    # 查询所有撤回申请
    requests = EvaluationRecord.query.filter_by(status='withdrawal_requested').options(
        joinedload(EvaluationRecord.evaluator),
        joinedload(EvaluationRecord.evaluatee),
        joinedload(EvaluationRecord.task)
    ).order_by(EvaluationRecord.updated_at.desc()).all()
    return render_template('admin/withdrawal_requests.html', requests=requests)


@app.route('/admin/evaluations/approve_withdrawal/<int:id>')
@login_required
@admin_required
def approve_withdrawal(id):
    from models import db, EvaluationRecord
    evaluation = EvaluationRecord.query.get_or_404(id)
    if evaluation.status == 'withdrawal_requested':
        # 将评估状态改为可编辑状态
        evaluation.status = 'returned'
        db.session.commit()
        flash('撤回申请已批准，评估记录已变为可编辑状态', 'success')
    else:
        flash('只能批准撤回申请状态的评估', 'warning')
    return redirect(url_for('withdrawal_requests'))


@app.route('/admin/evaluations/batch_approve_withdrawal/<int:task_id>')
@login_required
@admin_required
def batch_approve_withdrawal(task_id):
    from models import db, EvaluationRecord
    # 查询该任务下所有撤回申请状态的评估记录
    evaluations = EvaluationRecord.query.filter_by(
        task_id=task_id,
        status='withdrawal_requested'
    ).all()

    if not evaluations:
        flash('该任务下没有待批准的撤回申请', 'warning')
        return redirect(url_for('withdrawal_requests'))

    # 批量更新评估状态
    for evaluation in evaluations:
        evaluation.status = 'returned'
    db.session.commit()

    flash(f'成功批准 {len(evaluations)} 条撤回申请，相关评估记录已变为可编辑状态', 'success')
    return redirect(url_for('withdrawal_requests'))


@app.route('/admin/evaluations/batch_approve_withdrawal_by_evaluator/<int:task_id>/<int:evaluator_id>')
@login_required
@admin_required
def batch_approve_withdrawal_by_evaluator(task_id, evaluator_id):
    import logging
    from models import db, EvaluationRecord, Employee, EvaluationTask
    
    # 查询评估者和任务信息用于日志
    evaluator = Employee.query.get(evaluator_id)
    task = EvaluationTask.query.get(task_id)
    logging.info(f'管理员 {current_user.name} 尝试批量批准评估者 {evaluator.name if evaluator else evaluator_id} 在任务 {task.name if task else task_id} 下的撤回申请')

    # 查询该任务下指定评估者的所有撤回申请状态的评估记录
    evaluations = EvaluationRecord.query.filter_by(
        task_id=task_id,
        evaluator_id=evaluator_id,
        status='withdrawal_requested'
    ).all()

    if not evaluations:
        flash('该评估者在该任务下没有待批准的撤回申请', 'warning')
        logging.info(f'评估者 {evaluator.name if evaluator else evaluator_id} 在任务 {task.name if task else task_id} 下没有待批准的撤回申请')
        return redirect(url_for('withdrawal_requests'))

    # 批量更新评估状态
    evaluation_ids = [eval.id for eval in evaluations]
    for evaluation in evaluations:
        evaluation.status = 'returned'
    db.session.commit()

    flash(f'成功批准 {len(evaluations)} 条撤回申请，相关评估记录已变为可编辑状态', 'success')
    logging.info(f'成功批准评估者 {evaluator.name if evaluator else evaluator_id} 在任务 {task.name if task else task_id} 下的 {len(evaluations)} 条撤回申请，评估ID: {evaluation_ids}')
    return redirect(url_for('withdrawal_requests'))


@app.route('/admin/evaluations/batch_approve_all_withdrawals')
@login_required
@admin_required
def batch_approve_all_withdrawals():
    from models import db, EvaluationRecord
    # 查询所有撤回申请状态的评估记录
    evaluations = EvaluationRecord.query.filter_by(
        status='withdrawal_requested'
    ).all()

    if not evaluations:
        flash('没有待批准的撤回申请', 'warning')
        return redirect(url_for('withdrawal_requests'))

    # 批量更新评估状态
    for evaluation in evaluations:
        evaluation.status = 'returned'
    db.session.commit()

    flash(f'成功批准所有 {len(evaluations)} 条撤回申请，相关评估记录已变为可编辑状态', 'success')
    return redirect(url_for('withdrawal_requests'))


# 用户侧一键重新评估所有已退回评估
@app.route('/evaluations/batch_reassess_evaluations', methods=['POST'])
@login_required
def batch_reassess_evaluations():
    from models import db, EvaluationRecord
    # 获取当前用户作为评估者的所有已退回评估
    evaluations = EvaluationRecord.query.filter_by(
        evaluator_id=current_user.id,
        status='returned'
    ).all()

    if not evaluations:
        flash('没有可重新评估的已退回评估记录', 'warning')
        return redirect(url_for('my_evaluations'))

    # 批量更新评估状态为可编辑
    for evaluation in evaluations:
        evaluation.status = 'draft'
        evaluation.submitted_at = None
    db.session.commit()

    flash(f'成功将 {len(evaluations)} 条已退回评估记录变为可编辑状态', 'success')
    return redirect(url_for('my_evaluations'))


@app.route('/admin/evaluations/return/<int:id>')
@login_required
@admin_required
def return_evaluation(id):
    from models import db, EvaluationRecord
    evaluation = EvaluationRecord.query.options(joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)).get_or_404(id)
    if evaluation.status == 'submitted':
        evaluation.status = 'returned'
        db.session.commit()
        flash('评估已成功退回', 'success')
    else:
        flash('只能退回已提交的评估', 'warning')
    return redirect(url_for('admin_evaluation_query'))

# 员工数据导出功能
@logger.catch()
@handle_db_error
def validate_and_get_columns(headers):
    """验证表头并返回姓名、职位、密码列的索引"""
    required_headers = ['姓名', '职位', '默认密码']
    if not all(header in headers for header in required_headers):
        flash('Excel表头必须包含：姓名、职位、默认密码', 'danger')
        return None, None, None
    
    return (
        headers.index('姓名') + 1,
        headers.index('职位') + 1,
        headers.index('默认密码') + 1
    )


@app.route('/admin/employees/import', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_employee_import():
    from forms import EmployeeImportForm
    form = EmployeeImportForm()
    if not form.validate_on_submit():
        flash('表单验证失败: CSRF令牌缺失或无效', 'danger')
        return redirect(url_for('admin_employee_list'))
    if 'file' not in request.files:
        flash('未找到上传文件', 'danger')
    file = request.files['file']
    if file.filename == '':
        flash('未选择文件', 'danger')
    if not file.filename.endswith('.xlsx'):
        flash('请上传.xlsx格式的Excel文件', 'danger')
        return redirect(url_for('admin_employee_list'))
    
    from openpyxl import load_workbook
    from io import BytesIO
    file_content = file.read()
    wb = load_workbook(BytesIO(file_content))
    ws = wb.active
    
    # 验证表头并获取列索引
    try:
        headers = [cell.value for cell in ws[1]]
    except IndexError:
        flash('Excel文件格式错误，未找到表头行', 'danger')
        return redirect(url_for('admin_employee_list'))
    # 确保headers有效后再调用验证函数
    if headers:
        name_col, position_col, password_col = validate_and_get_columns(headers)
    else:
        flash('Excel文件格式错误，未找到表头', 'danger')
        return redirect(url_for('admin_employee_list'))
    if name_col is None:
        return redirect(url_for('admin_employee_list'))
        
    success_count = 0
    duplicate_names = []
    
    # 获取现有员工姓名列表
    existing_names = [emp.name for emp in Employee.query.all()]
        
    # 获取最大employee_id用于生成新编号
    max_emp = Employee.query.order_by(Employee.employee_id.desc()).first()
    start_id = 1
    if max_emp and max_emp.employee_id.isdigit():
        start_id = int(max_emp.employee_id) + 1
        
        # 从第二行开始读取数据
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[name_col - 1]
            position = row[position_col - 1]
            password = row[password_col - 1]
            
            if not all([name, position, password]):
                continue  # 跳过空行
            
            if name in existing_names or name in duplicate_names:
                if name not in duplicate_names:
                    duplicate_names.append(name)
                continue
            
            # 确保生成的employee_id唯一
            while Employee.query.filter_by(employee_id=str(start_id)).first():
                start_id += 1
            
            # 创建新员工
            new_employee = Employee(
                employee_id=str(start_id),
                name=name,
                position=position,
                is_admin=False
            )
            new_employee.set_password(password)
            
            db.session.add(new_employee)
            success_count += 1
            start_id += 1
            existing_names.append(name)  # 防止当前批次内重复

        db.session.commit()
        
        messages = []
        if success_count > 0:
            messages.append(f'成功导入 {success_count} 名员工')
        if duplicate_names:
            messages.append(f"跳过 {len(duplicate_names)} 个重复姓名: {', '.join(duplicate_names)}")
        
        if messages:
            for msg in messages:
                flash(msg, 'success' if '成功' in msg else 'warning')
        else:
            flash('未导入任何员工数据', 'info')
        
        return redirect(url_for('admin_employee_list'))



    
    # 查询所有员工数据
    employees = Employee.query.all()
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '员工信息'
    
    # 写入表头
    headers = ['ID', '员工编号', '姓名', '职位', '是否管理员', '创建时间', '更新时间']
    ws.append(headers)
    
    # 写入员工数据
    for emp in employees:
        ws.append([
            emp.id,
            emp.employee_id,
            emp.name,
            emp.position,
            '是' if emp.is_admin else '否',
            emp.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            emp.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    # 准备文件流
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成响应
    filename = f'员工信息导出_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8''{encoded_filename}'
    return response


# 批量退回评估者在特定任务的所有评估

@app.route('/admin/evaluations/batch_return/<int:evaluation_id>')
@login_required
@admin_required
@logger.catch()
def batch_return(evaluation_id):
    try:
        app.logger.info(f'批量退回请求: evaluation_id={evaluation_id}')
        from models import db, EvaluationRecord
        from sqlalchemy import func
        # 通过单个评估记录获取任务和评估者信息
        ref_eval = EvaluationRecord.query.get_or_404(evaluation_id)
        app.logger.info(f'批量退回: 获取参考评估记录 - task_id={ref_eval.task_id}, evaluator_id={ref_eval.evaluator_id}')
        evaluations = EvaluationRecord.query.filter(
            EvaluationRecord.task_id == ref_eval.task_id,
            EvaluationRecord.evaluator_id == ref_eval.evaluator_id,
            func.lower(EvaluationRecord.status) == 'submitted'
        ).all()
        app.logger.info(f'批量退回: 查询到{len(evaluations)}条符合条件的评估记录')
        
        if evaluations:
            for eval in evaluations:
                eval.status = 'returned'
            try:
                db.session.commit()
                app.logger.info(f'批量退回: 成功提交{len(evaluations)}条记录的状态更新')
                flash(f'成功退回 {len(evaluations)} 条评估记录', 'success')
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'批量退回: 数据库提交失败 - {str(e)}', exc_info=True)
                flash(f'退回失败: {str(e)}', 'danger')
        else:
            flash('没有找到可退回的评估记录', 'warning')
    except Exception as e:
        app.logger.error(f'批量退回操作发生异常: {str(e)}', exc_info=True)
        flash(f'操作失败: {str(e)}', 'danger')
        return redirect(url_for('admin_evaluation_query'))
    return redirect(url_for('admin_evaluation_query'))

@app.route('/admin/evaluations/delete/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def delete_evaluation(id):
    evaluation = EvaluationRecord.query.options(joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)).get_or_404(id)
    db.session.delete(evaluation)
    db.session.commit()
    flash('评估记录已成功删除', 'success')
    return redirect(url_for('admin_evaluation_query'))

# 初始化数据库
@app.before_first_request
def create_tables():
    from models import EvaluationDimension
    db.create_all()
    # 检查是否已有默认维度，没有则添加
    if not EvaluationDimension.query.first():
        default_dimensions = [
            ('本职工作', 0.5),
            ('附加业绩', 0.2),
            ('合规守纪', 0.15),
            ('互助', 0.1),
            ('勤勉', 0.05)
        ]
        for name, weight in default_dimensions:
            dimension = EvaluationDimension(name=name, weight=weight)
            db.session.add(dimension)
        db.session.commit()

@app.route('/batch_evaluate')
def batch_evaluate():
    # 检查是否为管理员
    if current_user.employee_id == '10000':
        flash('管理员账号不能参与评估打分', 'danger')
        return redirect(url_for('index'))
    evaluator_id = request.args.get('evaluator_id')
    task_id = request.args.get('task_id')
    
    if not evaluator_id:
        flash('请选择评估者', 'warning')
        return redirect(url_for('index'))
    
    if not task_id:
        flash('请选择评估任务', 'warning')
        return redirect(url_for('evaluate_page'))
    
    # 获取评估任务
    task = EvaluationTask.query.get(task_id)
    if not task:
        flash('评估任务不存在', 'danger')
        return redirect(url_for('evaluate_page'))
    
    # 检查是否已提交该任务的评估
    existing_submission = EvaluationRecord.query.filter_by(
        evaluator_id=evaluator_id,
        task_id=task_id,
        status='submitted'
    ).options(joinedload(EvaluationRecord.scores).joinedload(EvaluationScore.dimension)).first()
    
    if existing_submission:
        flash('你已提交本次任务，若需修改请通知管理员退回。', 'warning')
        return redirect(url_for('evaluate_page'))
    
    # 检查任务是否存在待审核的撤回申请
    has_pending_withdrawal = EvaluationRecord.query.filter_by(
        evaluator_id=evaluator_id,
        task_id=task_id,
        status='withdrawal_requested'
    ).first() is not None
    
    # 获取所有被评估者（排除自己、管理员、被冻结的员工以及ID小于2的员工）
    evaluatees = Employee.query.filter(
        Employee.id != evaluator_id,
        Employee.employee_id != '10000',
        Employee.employee_id < '2',
        Employee.is_frozen == False
    ).all()
    # 获取所有评估维度
    dimensions = EvaluationDimension.query.all()
    
    # 获取当前任务
    task = EvaluationTask.query.get(task_id)
    return render_template('evaluation/batch_evaluation.html', 
                          evaluator_id=evaluator_id, 
                          task_id=task_id, 
                          task=task, 
                          evaluatees=evaluatees, 
                          dimensions=dimensions, 
                          has_pending_withdrawal=has_pending_withdrawal)

if __name__ == '__main__':
    with app.app_context():
        try:
              import os
              print(f"当前工作目录: {os.getcwd()}", flush=True)
              print(f"数据库URI: {app.config['SQLALCHEMY_DATABASE_URI']}", flush=True)
              print("导入模型前元数据 tables: {db.metadata.tables.keys()}", flush=True)
              print("开始创建数据库表...")
              from models import Employee, EvaluationDimension, EvaluationScore, EvaluationRecord
              print("导入模型后元数据 tables: {db.metadata.tables.keys()}", flush=True)
              db.create_all()
              # 验证表是否创建成功
              tables = db.engine.table_names()
              print(f"创建的数据库表: {tables}", flush=True)
              print("数据库表创建成功", flush=True)
              
              # 检查是否已有默认维度，没有则添加
              print("检查默认维度...")
              if not EvaluationDimension.query.first():
                  print("添加默认维度...")
                  default_dimensions = [
                      ('本职工作', 0.5),
                      ('附加业绩', 0.2),
                      ('合规守纪', 0.15),
                      ('互助', 0.1),
                      ('勤勉', 0.05)
                  ]
                  for name, weight in default_dimensions:
                      dimension = EvaluationDimension(name=name, weight=weight)
                      db.session.add(dimension)
                  db.session.commit()
                  print("默认维度添加成功")
              else:
                  print("默认维度已存在")
        except Exception as e:
            print(f"数据库初始化错误: {str(e)}")
            db.session.rollback()
        finally:
            print("数据库初始化流程结束")

app.run(debug=True, host='0.0.0.0', port=3008, use_reloader=False)


